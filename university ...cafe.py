import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import uuid
import json

# =================================================================================================
# GLOBAL VARIABLES AND FILE PATHS
# =================================================================================================

excel_file = 'cafe_orders_summary.xlsx'
summary_headers = ["Order ID", "Date", "Menu Type", "Items", "Total Price"]

settings_file = 'cafe_settings.json'
breakfast_menu_file = 'menus/breakfast_menu.json'
lunch_menu_file = 'menus/lunch_menu.json'

# Ensure 'menus' directory exists for storing menu JSON files
os.makedirs('menus', exist_ok=True)

# Default menus if JSON files don't exist (used for initial creation)
default_breakfast_menu = {
    "Lahori Channy": {"default": 140},
    "Aloo Bhujia": {"default": 150},
    "Paratha": {"default": 50},
    "Omellete": {"default": 50},
    "Fried Egg": {"default": 50},
    "Boiled Egg": {"default": 50},
    "French Toast": {"default": 120},
    "Simple Toast": {"default": 50},
    "Suji Halwa": {"default": 150},
    "Green Tea": {"Small": 50, "Large": 80},
    "Tea": {"Small": 50, "Large": 80},
    "Roghni Naan": {"default": 50},
    "Plain Naan": {"default": 20}
}

default_lunch_menu = {
    "Chicken Biryani": {"default": 230},
    "Channa Pulao": {"default": 180},
    "Chicken Piece": {"default": 100},
    "Chicken Qorma": {"default": 220},
    "White Qorma": {"default": 250},
    "Daal Mash": {"default": 140},
    "Mix Vegetable": {"default": 140},
    "Macroni": {"default": 290},
    "Chowmein": {"default": 290},
    "Daleem": {"default": 160},
    "Daleem With Rice": {"default": 140},
    "Shami Kabab": {"default": 60},
    "Raita": {"default": 20},
    "Salad": {"default": 50},
    "Roti": {"default": 16},
    "Naan": {"default": 20},
    "Packing Box": {"default": 30},
    "Paper Cup": {"default": 10},
    "Kabli Pulao":{"default":45},
}

# Variables to hold Tkinter widgets/vars that need to be globally accessible
# or accessible across different functions when the menu management window is open.
# Initialized to None or empty as they are created when manage_menu_items is called.
menu_manage_window = None
item_name_entry = None
default_price_entry = None
has_options_var = None
option_input_frame = None
option_entries_widgets = [] # List to hold (label, entry) tuples for options
option_vars_list = []       # List to hold StringVar for option names and prices
menu_tree = None # Treeview widget for menu management

# =================================================================================================
# MENU LOADING/SAVING FUNCTIONS
# =================================================================================================

def load_menu_from_file(file_path, default_menu):
    """Loads a menu from a JSON file, or uses a default if file not found/corrupted."""
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as f:
                loaded_menu = json.load(f)
                if not isinstance(loaded_menu, dict):
                    raise ValueError("Menu file content is not a dictionary.")
                return loaded_menu
        except (json.JSONDecodeError, ValueError) as e:
            messagebox.showwarning("Menu Load Error", f"Could not read {os.path.basename(file_path)}: {e}\nUsing default menu and recreating file.")
            save_menu_to_file(file_path, default_menu) # Overwrite corrupted file
            return default_menu
    else:
        save_menu_to_file(file_path, default_menu)
        return default_menu

def save_menu_to_file(file_path, menu_data):
    """Saves a menu dictionary to a JSON file."""
    try:
        with open(file_path, 'w') as f:
            json.dump(menu_data, f, indent=4)
    except Exception as e:
        messagebox.showerror("Menu Save Error", f"Could not save menu to {os.path.basename(file_path)}: {e}")

# Initial load of menus at application start
current_breakfast_menu = load_menu_from_file(breakfast_menu_file, default_breakfast_menu)
current_lunch_menu = load_menu_from_file(lunch_menu_file, default_lunch_menu)

# This dictionary maps menu names (strings) to the actual menu data dictionaries
all_menus_loaded = {
    "Breakfast": current_breakfast_menu,
    "Lunch": current_lunch_menu
}

# =================================================================================================
# PERSISTENT SETTINGS
# =================================================================================================

default_settings = {
    "cafe_name": "COMSATS University Islamabad Café",
    "current_menu": "Breakfast" # Initial menu type
}

def load_settings():
    """Loads cafe settings from a JSON file."""
    if os.path.exists(settings_file):
        try:
            with open(settings_file, 'r') as f:
                settings = json.load(f)
                for key, default_val in default_settings.items():
                    if key not in settings:
                        settings[key] = default_val
                return settings
        except json.JSONDecodeError as e:
            messagebox.showwarning("Settings Load Error", f"Could not read {settings_file}: {e}\nUsing default settings and recreating file.")
            save_settings(default_settings)
            return default_settings
    save_settings(default_settings)
    return default_settings

def save_settings(settings):
    """Saves cafe settings to a JSON file."""
    try:
        with open(settings_file, 'w') as f:
            json.dump(settings, f, indent=4)
    except Exception as e:
        messagebox.showerror("Settings Save Error", f"Could not save settings to {settings_file}: {e}")

cafe_settings = load_settings()

def get_current_menu_data():
    """Returns the currently selected menu dictionary based on cafe_settings."""
    return all_menus_loaded.get(cafe_settings["current_menu"], all_menus_loaded["Breakfast"])

# Ensure Excel file exists with headers
if not os.path.exists(excel_file):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cafe Orders"
        ws.append(summary_headers)
        wb.save(excel_file)
    except Exception as e:
        messagebox.showerror("File Error", f"Could not create Excel file '{excel_file}': {e}")

# =================================================================================================
# GUI SETUP (MAIN WINDOW)
# =================================================================================================

root = tk.Tk()
root.title(f"☕ {cafe_settings['cafe_name']} - {cafe_settings['current_menu']} Order App")
root.geometry("680x850")
root.configure(bg="#f4f4f4")

# Apply modern theme for ttk widgets
style = ttk.Style()
style.theme_use('clam')
style.configure('TFrame', background='#f4f4f4')
style.configure('TLabel', background='#f4f4f4', foreground='#333')
style.configure('TButton', font=('Arial', 10), padding=5)
style.map('TButton',
          foreground=[('active', 'white')],
          background=[('active', '#F57C00')])
style.configure('TCombobox', fieldbackground='white', background='white', font=('Arial', 10))
style.configure('Treeview.Heading', font=('Arial', 10, 'bold'), background='#007bff', foreground='white')
style.configure('Treeview', font=('Arial', 9), rowheight=25)


# Header Frame
header_frame = tk.Frame(root, bg="#f4f4f4")
header_frame.pack(pady=10, fill='x')

cafe_name_label = tk.Label(header_frame, text=f"☕ {cafe_settings['cafe_name']}", font=('Arial', 24, 'bold'), bg="#f4f4f4", fg="#333")
cafe_name_label.pack()
menu_type_label = tk.Label(header_frame, text=f"{cafe_settings['current_menu']} Ordering System", font=('Arial', 16), bg="#f4f4f4", fg="#555")
menu_type_label.pack()

# ========= Receipt Display - MOVED UP FOR DEFINITION SAFETY ==========
receipt_box = tk.Text(root, height=10, width=70, bg="white", font=('Courier New', 10))
receipt_box.pack(pady=10)

# ========= Scrollable Menu Area ==========
canvas_frame = tk.Frame(root)
canvas_frame.pack(pady=10, fill='both', expand=True)

canvas = tk.Canvas(canvas_frame, height=350, bg="#ffffff", bd=0, highlightthickness=0)
scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas, bg="#ffffff")

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas_window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

def on_canvas_resize(event):
    """Adjusts the width of the scrollable frame within the canvas."""
    canvas.itemconfig(canvas_window_id, width=event.width)

canvas.bind("<Configure>", on_canvas_resize)

item_data = {} # Stores IntVar for quantity and StringVar for options for each item on the main menu

def update_menu_display():
    """Renders or re-renders the entire menu display on the main window
    based on the currently selected menu (Breakfast/Lunch)."""
    # Clear existing items from the scrollable frame
    for widget in scrollable_frame.winfo_children():
        widget.destroy()

    # Clear item_data to remove quantities from previous menu
    item_data.clear()

    # Configure grid columns for item layout
    scrollable_frame.grid_columnconfigure(0, weight=1, minsize=200) # Item Name
    scrollable_frame.grid_columnconfigure(1, weight=0, minsize=120) # Options Dropdown
    scrollable_frame.grid_columnconfigure(2, weight=0, minsize=30)  # Minus Button
    scrollable_frame.grid_columnconfigure(3, weight=0, minsize=40)  # Quantity Label
    scrollable_frame.grid_columnconfigure(4, weight=0, minsize=30)  # Plus Button

    current_menu = get_current_menu_data() # Get the active menu data (e.g., breakfast_menu or lunch_menu)
    row_idx = 0
    for item, options in current_menu.items():
        item_frame = tk.Frame(scrollable_frame, bg="#ffffff", pady=5)
        item_frame.grid(row=row_idx, column=0, columnspan=5, sticky='ew', padx=10)

        # Determine price display (e.g., "(150 PKR)" or "(50-80 PKR)")
        price_display = ""
        if "default" in options:
            price_display = f"({options['default']} PKR)"
        else:
            prices = list(options.values())
            if prices:
                price_display = f"({min(prices)}-{max(prices)} PKR)"

        tk.Label(item_frame, text=f"{item} {price_display}", font=('Arial', 12), anchor="w", bg="#ffffff").grid(row=0, column=0, sticky='ew', padx=(0, 5))

        # Setup quantity and option variables for each item
        qty_var = tk.IntVar(value=0)
        option_var = tk.StringVar(value="")
        item_data[item] = {"qty": qty_var, "option": option_var, "option_dropdown": None}

        # If item has multiple options, create a Combobox
        if "default" not in options:
            option_values = list(options.keys())
            if option_values:
                option_var.set(option_values[0]) # Set default to first available option
                option_dropdown = ttk.Combobox(item_frame, textvariable=option_var, values=option_values, state="readonly", width=12)
                option_dropdown.grid(row=0, column=1, padx=5, sticky='ew')
                item_data[item]["option_dropdown"] = option_dropdown
        else:
            # Placeholder for items with default prices (no dropdown needed)
            tk.Label(item_frame, text="", width=12, bg="#ffffff").grid(row=0, column=1, padx=5, sticky='ew')

        # Quantity control buttons and label
        tk.Button(item_frame, text="-", command=lambda i=item: decrease(i), width=2, font=('Arial', 10)).grid(row=0, column=2, padx=(5,0))
        tk.Label(item_frame, textvariable=qty_var, width=3, relief="solid", bg="#fff", font=('Arial', 12)).grid(row=0, column=3, padx=0)
        tk.Button(item_frame, text="+", command=lambda i=item: increase(i), width=2, font=('Arial', 10)).grid(row=0, column=4, padx=(0,5))
        
        # Configure column weights within the item frame for proper resizing
        item_frame.grid_columnconfigure(0, weight=1)
        item_frame.grid_columnconfigure(1, weight=0)
        item_frame.grid_columnconfigure(2, weight=0)
        item_frame.grid_columnconfigure(3, weight=0)
        item_frame.grid_columnconfigure(4, weight=0)

        row_idx += 1
    
    # Clear the receipt box when the menu display is updated (e.g., menu type changed)
    receipt_box.delete("1.0", tk.END)

# Initial display of menu when the application starts
update_menu_display()

def increase(item):
    """Increases the quantity of a selected item.
    If it's an item with options, ensures an option is selected on first increase."""
    current_qty = item_data[item]["qty"].get()
    item_data[item]["qty"].set(current_qty + 1)
    # If adding first item and it has options, ensure an option is pre-selected
    if current_qty == 0 and "default" not in get_current_menu_data()[item] and not item_data[item]["option"].get():
        option_values = list(get_current_menu_data()[item].keys())
        if option_values:
            item_data[item]["option"].set(option_values[0])

def decrease(item):
    """Decreases the quantity of a selected item, ensuring it doesn't go below zero."""
    qty = item_data[item]["qty"].get()
    if qty > 0:
        item_data[item]["qty"].set(qty - 1)

# ========= Generate Bill Function ==========
def generate_bill():
    """Calculates the total bill, displays it in the receipt box, and saves the order to Excel."""
    receipt_box.delete("1.0", tk.END) # Clear previous receipt

    total = 0
    selected_items_for_excel = [] # List to store items for Excel summary
    item_string_list_for_excel = [] # Formatted strings for Excel "Items" column

    order_id = str(uuid.uuid4())[:8].upper() # Generate a short unique ID
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    menu_type_recorded = cafe_settings['current_menu'] # Record which menu type this order came from

    # Construct receipt header
    receipt_box.insert(tk.END, f"       ☕ {cafe_settings['cafe_name']} Receipt\n")
    receipt_box.insert(tk.END, "=================================================\n")
    receipt_box.insert(tk.END, f"Order ID: {order_id}\n")
    receipt_box.insert(tk.END, f"Date: {date_str}\n")
    receipt_box.insert(tk.END, f"Menu Type: {menu_type_recorded}\n") # Display menu type on receipt
    receipt_box.insert(tk.END, "-------------------------------------------------\n")
    receipt_box.insert(tk.END, f"{'Item':25} {'Qty':>3} {'Unit Price':>12} {'Total Price':>12}\n")
    receipt_box.insert(tk.END, "-------------------------------------------------\n")

    current_active_menu = get_current_menu_data() # Get the current menu in use (live data)

    any_item_selected = False
    for item_name_key, data in item_data.items(): # Iterate through all items defined in item_data
        qty = data["qty"].get()
        if qty > 0:
            any_item_selected = True
            selected_option = data["option"].get()
            
            # Ensure the item is still present in the current active menu
            if item_name_key not in current_active_menu:
                messagebox.showerror("Menu Error", f"'{item_name_key}' is no longer available in the current {menu_type_recorded} menu. Please reset and try again.")
                receipt_box.delete("1.0", tk.END)
                return

            item_options_prices = current_active_menu[item_name_key]
            unit_price = 0
            display_item_name = ""
            
            if "default" in item_options_prices: # Item has a single default price
                unit_price = item_options_prices["default"]
                display_item_name = item_name_key
            elif selected_option and selected_option in item_options_prices: # Item has options and one is selected
                unit_price = item_options_prices[selected_option]
                display_item_name = f"{item_name_key} ({selected_option})"
            else: # Item has options but none selected (shouldn't happen if increase() works correctly)
                messagebox.showerror("Selection Error", f"Please select an option for '{item_name_key}'.")
                receipt_box.delete("1.0", tk.END)
                return

            item_total_price = qty * unit_price
            total += item_total_price
            selected_items_for_excel.append((display_item_name, qty, unit_price, item_total_price))
            item_string_list_for_excel.append(f"{display_item_name} x{qty}") # Format for Excel's 'Items' column

            receipt_box.insert(tk.END, f"{display_item_name:25} {qty:>3} {unit_price:>12} {item_total_price:>12} PKR\n")

    if not any_item_selected:
        messagebox.showwarning("Empty Order", "Please select at least one item to generate a bill.")
        return

    # Construct receipt footer
    receipt_box.insert(tk.END, "-------------------------------------------------\n")
    receipt_box.insert(tk.END, f"{'TOTAL':>43} {total:>12} PKR\n")
    receipt_box.insert(tk.END, "=================================================\n")

    # Save order to Excel file
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        # Append data including the menu type
        ws.append([order_id, date_str, menu_type_recorded, ", ".join(item_string_list_for_excel), total])
        wb.save(excel_file)
        messagebox.showinfo("Bill Generated", "Order saved successfully!")
    except Exception as e:
        messagebox.showerror("File Save Error", f"Could not save order to Excel: {e}\n"
                                               "Please ensure the Excel file is closed and not corrupted.")


# ========= Reset Function ==========
def reset_form():
    """Resets all item quantities to 0, clears selected options, and clears the receipt box."""
    items_to_reset = list(item_data.keys()) 
    for item in items_to_reset:
        if item in item_data:
            item_data[item]["qty"].set(0)
            
            current_menu = get_current_menu_data()
            if item in current_menu and "default" not in current_menu[item]:
                if item_data[item]["option_dropdown"]: # Ensure dropdown exists
                    option_values = list(current_menu[item].keys())
                    if option_values:
                        item_data[item]["option"].set(option_values[0])
                    else:
                        item_data[item]["option"].set("")
    receipt_box.delete("1.0", tk.END)


# ========= Sales Report Window ==========
def open_sales_report():
    """Opens a new Toplevel window to display past orders from the Excel file."""
    report_window = tk.Toplevel(root)
    report_window.title("📊 Sales Report")
    report_window.geometry("800x600")
    report_window.transient(root)
    report_window.grab_set()

    report_frame = ttk.Frame(report_window, padding="10 10 10 10")
    report_frame.pack(fill="both", expand=True)

    report_search_frame = ttk.Frame(report_frame)
    report_search_frame.pack(pady=5, fill='x')
    ttk.Label(report_search_frame, text="Search Order/Item/Menu Type:", font=('Arial', 10)).pack(side='left', padx=5)
    report_search_entry = ttk.Entry(report_search_frame, width=40, font=('Arial', 10))
    report_search_entry.pack(side='left', fill='x', expand=True)

    tree_frame = ttk.Frame(report_frame)
    tree_frame.pack(fill="both", expand=True, pady=10)

    tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
    tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")

    columns = summary_headers
    tree = ttk.Treeview(tree_frame, columns=columns, show="headings",
                        yscrollcommand=tree_scroll_y.set,
                        xscrollcommand=tree_scroll_x.set)

    tree_scroll_y.config(command=tree.yview)
    tree_scroll_x.config(command=tree.xview)

    tree_scroll_y.pack(side="right", fill="y")
    tree_scroll_x.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    for col in columns:
        tree.heading(col, text=col, anchor="w")
        tree.column(col, width=100, minwidth=80)

    tree.column("Order ID", width=80)
    tree.column("Date", width=150)
    tree.column("Menu Type", width=100)
    tree.column("Items", width=250)
    tree.column("Total Price", width=100, anchor="e")

    def load_report_data(search_term=""):
        """Loads data from Excel into the Treeview, with optional search filtering."""
        for i in tree.get_children():
            tree.delete(i)
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            for row_index, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if search_term.lower() in str(row_data).lower() or not search_term:
                    tree.insert("", "end", values=row_data)
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"Sales report file '{excel_file}' not found.")
            report_window.destroy()
        except Exception as e:
            messagebox.showerror("File Read Error", f"Could not read sales report: {e}\n"
                                                   "Please ensure the Excel file is closed and not corrupted.")
            report_window.destroy()

    report_search_entry.bind("<KeyRelease>", lambda e: load_report_data(report_search_entry.get()))
    load_report_data()

    ttk.Button(report_frame, text="Close", command=report_window.destroy).pack(pady=10)


# ========= Manage Sales Data Window ==========
def manage_sales_data():
    """Opens a window to manage sales data (clear/backup/restore)."""
    sales_window = tk.Toplevel(root)
    sales_window.title("🗑️ Manage Sales Data")
    sales_window.geometry("400x300")
    sales_window.transient(root)
    sales_window.grab_set()

    frame = ttk.Frame(sales_window, padding="20")
    frame.pack(fill="both", expand=True)

    def clear_sales_data():
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all sales data? This cannot be undone!"):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Cafe Orders"
                ws.append(summary_headers)
                wb.save(excel_file)
                messagebox.showinfo("Success", "Sales data has been cleared successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Could not clear sales data: {e}")

    ttk.Label(frame, text="Sales Data Management", font=('Arial', 14, 'bold')).pack(pady=10)
    ttk.Button(frame, text="Clear All Sales Data", command=clear_sales_data).pack(pady=20)
    ttk.Button(frame, text="Close", command=sales_window.destroy).pack(pady=10)

def manage_menu_items():
    """Opens a window to manage (add/edit/delete) menu items."""
    global menu_manage_window, item_name_entry, default_price_entry, menu_tree, current_breakfast_menu, current_lunch_menu
    
    menu_manage_window = tk.Toplevel(root)
    menu_manage_window.title("🍽️ Manage Menu Items")
    menu_manage_window.geometry("600x500")
    menu_manage_window.transient(root)

    # Create frames
    input_frame = ttk.Frame(menu_manage_window, padding="10")
    input_frame.pack(fill='x')

    # Menu type selection
    ttk.Label(input_frame, text="Menu Type:").grid(row=0, column=0, padx=5, pady=5)
    menu_type_var = tk.StringVar(value="Breakfast")
    menu_type_combo = ttk.Combobox(input_frame, textvariable=menu_type_var, values=["Breakfast", "Lunch"], state="readonly")
    menu_type_combo.grid(row=0, column=1, padx=5, pady=5)

    # Item name input
    ttk.Label(input_frame, text="Item Name:").grid(row=1, column=0, padx=5, pady=5)
    item_name_entry = ttk.Entry(input_frame)
    item_name_entry.grid(row=1, column=1, padx=5, pady=5)

    # Price input
    ttk.Label(input_frame, text="Price (PKR):").grid(row=2, column=0, padx=5, pady=5)
    default_price_entry = ttk.Entry(input_frame)
    default_price_entry.grid(row=2, column=1, padx=5, pady=5)

    # Create Treeview to display menu items
    tree_frame = ttk.Frame(menu_manage_window)
    tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

    menu_tree = ttk.Treeview(tree_frame, columns=("Item", "Price"), show="headings")
    menu_tree.heading("Item", text="Item Name")
    menu_tree.heading("Price", text="Price (PKR)")
    menu_tree.pack(fill='both', expand=True)

    def refresh_main_menu():
        """Refreshes the main menu display and resets quantities"""
        reset_form()  # Reset all quantities
        update_menu_display()  # Update the main menu display

    def load_menu_items():
        """Loads menu items into the treeview based on selected menu type"""
        menu_tree.delete(*menu_tree.get_children())
        selected_menu = menu_type_var.get()
        menu_data = all_menus_loaded[selected_menu]
        for item, price_data in menu_data.items():
            price = price_data.get("default", "N/A")
            menu_tree.insert("", "end", values=(item, price))

    def save_menu_item():
        """Saves a new menu item or updates an existing one"""
        item_name = item_name_entry.get().strip()
        price = default_price_entry.get().strip()
        
        if not item_name or not price:
            messagebox.showwarning("Input Error", "Please fill in both item name and price.")
            return
        
        try:
            price = float(price)
        except ValueError:
            messagebox.showwarning("Input Error", "Price must be a number.")
            return

        selected_menu = menu_type_var.get()
        if selected_menu == "Breakfast":
            current_breakfast_menu[item_name] = {"default": price}
            save_menu_to_file(breakfast_menu_file, current_breakfast_menu)
            all_menus_loaded["Breakfast"] = current_breakfast_menu
        else:
            current_lunch_menu[item_name] = {"default": price}
            save_menu_to_file(lunch_menu_file, current_lunch_menu)
            all_menus_loaded["Lunch"] = current_lunch_menu
        
        load_menu_items()
        refresh_main_menu()  # Refresh the main menu display
        item_name_entry.delete(0, tk.END)
        default_price_entry.delete(0, tk.END)
        messagebox.showinfo("Success", f"Menu item '{item_name}' has been saved.")

    def delete_menu_item():
        """Deletes the selected menu item"""
        selected_item = menu_tree.selection()
        if not selected_item:
            messagebox.showwarning("Selection Error", "Please select an item to delete.")
            return

        item_name = menu_tree.item(selected_item[0])['values'][0]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{item_name}'?"):
            selected_menu = menu_type_var.get()
            if selected_menu == "Breakfast":
                if item_name in current_breakfast_menu:
                    del current_breakfast_menu[item_name]
                    save_menu_to_file(breakfast_menu_file, current_breakfast_menu)
                    all_menus_loaded["Breakfast"] = current_breakfast_menu
            else:
                if item_name in current_lunch_menu:
                    del current_lunch_menu[item_name]
                    save_menu_to_file(lunch_menu_file, current_lunch_menu)
                    all_menus_loaded["Lunch"] = current_lunch_menu
            
            load_menu_items()
            refresh_main_menu()  # Refresh the main menu display
            messagebox.showinfo("Success", f"Menu item '{item_name}' has been deleted.")

    # Allow editing by double-clicking on a tree item
    def on_tree_double_click(event):
        item = menu_tree.selection()[0]
        if item:
            values = menu_tree.item(item)['values']
            item_name_entry.delete(0, tk.END)
            item_name_entry.insert(0, values[0])
            default_price_entry.delete(0, tk.END)
            default_price_entry.insert(0, values[1])

    menu_tree.bind('<Double-1>', on_tree_double_click)

    # Buttons frame
    button_frame = ttk.Frame(menu_manage_window)
    button_frame.pack(fill='x', padx=10, pady=10)

    ttk.Button(button_frame, text="Save Item", command=save_menu_item).pack(side='left', padx=5)
    ttk.Button(button_frame, text="Delete Item", command=delete_menu_item).pack(side='left', padx=5)
    ttk.Button(button_frame, text="Close", command=menu_manage_window.destroy).pack(side='right', padx=5)

    menu_type_combo.bind('<<ComboboxSelected>>', lambda e: load_menu_items())
    load_menu_items()

# =================================================================================================
# SETTINGS WINDOW
# =================================================================================================

def open_settings():
    """Opens a new Toplevel window for application settings."""
    settings_window = tk.Toplevel(root)
    settings_window.title("⚙️ Settings")
    settings_window.geometry("400x350")
    settings_window.transient(root)
    settings_window.grab_set()

    settings_frame = ttk.Frame(settings_window, padding="15")
    settings_frame.pack(fill="both", expand=True)

    # Cafe Name Setting
    ttk.Label(settings_frame, text="Cafe Name:", font=('Arial', 12)).grid(row=0, column=0, padx=5, pady=5, sticky='w')
    cafe_name_entry = ttk.Entry(settings_frame, width=30, font=('Arial', 12))
    cafe_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
    cafe_name_entry.insert(0, cafe_settings['cafe_name'])

    # Menu Type Selection
    ttk.Label(settings_frame, text="Current Menu:", font=('Arial', 12)).grid(row=1, column=0, padx=5, pady=5, sticky='w')
    
    menu_type_var = tk.StringVar(value=cafe_settings['current_menu'])
    menu_type_dropdown = ttk.Combobox(settings_frame, textvariable=menu_type_var, 
                                      values=list(all_menus_loaded.keys()), state="readonly", width=27, font=('Arial', 12))
    menu_type_dropdown.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

    def save_and_close_settings():
        """Saves updated cafe name and menu type, then closes settings window."""
        new_cafe_name = cafe_name_entry.get().strip()
        new_menu_type = menu_type_var.get().strip()

        if not new_cafe_name:
            messagebox.showwarning("Input Error", "Cafe Name cannot be empty.")
            return
        if not new_menu_type:
            messagebox.showwarning("Input Error", "Menu Type cannot be empty.")
            return

        menu_type_changed = (cafe_settings['current_menu'] != new_menu_type)

        cafe_settings['cafe_name'] = new_cafe_name
        cafe_settings['current_menu'] = new_menu_type
        save_settings(cafe_settings) # Saves the overall application settings

        cafe_name_label.config(text=f"☕ {cafe_settings['cafe_name']}")
        root.title(f"☕ {cafe_settings['cafe_name']} - {cafe_settings['current_menu']} Order App")
        menu_type_label.config(text=f"{cafe_settings['current_menu']} Ordering System")

        if menu_type_changed:
            reset_form() # Reset quantities when menu type changes
            update_menu_display() # Re-render the menu
            messagebox.showinfo("Settings Saved", "Settings updated successfully! Menu has been switched.")
        else:
            messagebox.showinfo("Settings Saved", "Settings updated successfully!")

        settings_window.destroy()

    ttk.Button(settings_frame, text="Save Settings", command=save_and_close_settings).grid(row=2, column=0, pady=10, padx=5, sticky='ew')
    ttk.Button(settings_frame, text="Cancel", command=settings_window.destroy).grid(row=2, column=1, pady=10, padx=5, sticky='ew')

    # New Button for Managing Sales Data
    ttk.Button(settings_frame, text="🗑️ Manage Sales Data", command=manage_sales_data,
               style='Manage.TButton').grid(row=3, column=0, columnspan=2, pady=10, padx=5, sticky='ew')
    
    # New Button for Managing Menu Items
    ttk.Button(settings_frame, text="🍽️ Manage Menu Items", command=manage_menu_items,
               style='Manage.TButton').grid(row=4, column=0, columnspan=2, pady=5, padx=5, sticky='ew')

    style.configure('Manage.TButton', background='#17a2b8', foreground='white', font=('Arial', 10, 'bold'))
    style.map('Manage.TButton',
              foreground=[('active', 'white')],
              background=[('active', '#138496')])


# =================================================================================================
# MAIN WINDOW CONTROL BUTTONS
# =================================================================================================

button_frame = tk.Frame(root, bg="#f4f4f4")
button_frame.pack(pady=(10, 20), fill='x')

generate_btn = tk.Button(
    button_frame,
    text="🧾 GENERATE BILL",
    command=generate_bill,
    font=('Arial', 16, 'bold'),
    bg="#FF9800",
    fg="white",
    activebackground="#F57C00",
    activeforeground="white",
    relief="raised",
    bd=6,
    padx=20,
    pady=10,
    cursor="hand2"
)
generate_btn.pack(side='left', expand=True, padx=10)

def on_enter_generate(e): generate_btn.config(bg="#F57C00")
def on_leave_generate(e): generate_btn.config(bg="#FF9800")
generate_btn.bind("<Enter>", on_enter_generate)
generate_btn.bind("<Leave>", on_leave_generate)

next_btn = tk.Button(
    button_frame,
    text="⏭️ NEXT ORDER",
    command=reset_form,
    font=('Arial', 14, 'bold'),
    bg="#9C27B0",
    fg="white",
    activebackground="#7B1FA2",
    activeforeground="white",
    relief="raised",
    bd=5,
    padx=20,
    pady=8,
    cursor="hand2"
)
next_btn.pack(side='left', expand=True, padx=10)

def on_enter_next(e): next_btn.config(bg="#7B1FA2")
def on_next_leave(e): next_btn.config(bg="#9C27B0")
next_btn.bind("<Enter>", on_enter_next)
next_btn.bind("<Leave>", on_next_leave)

report_btn = tk.Button(
    button_frame,
    text="📊 SALES REPORT",
    command=open_sales_report,
    font=('Arial', 14, 'bold'),
    bg="#007bff",
    fg="white",
    activebackground="#0056b3",
    activeforeground="white",
    relief="raised",
    bd=5,
    padx=20,
    pady=8,
    cursor="hand2"
)
report_btn.pack(side='left', expand=True, padx=10)

def on_enter_report(e): report_btn.config(bg="#0056b3")
def on_leave_report(e): report_btn.config(bg="#007bff")
report_btn.bind("<Enter>", on_enter_report)
report_btn.bind("<Leave>", on_leave_report)

settings_btn = tk.Button(
    root,
    text="⚙️ SETTINGS",
    command=open_settings,
    font=('Arial', 10),
    bg="#6c757d",
    fg="white",
    activebackground="#5a6268",
    activeforeground="white",
    relief="raised",
    bd=3,
    padx=10,
    pady=5,
    cursor="hand2"
)
settings_btn.pack(side='bottom', anchor='se', padx=15, pady=10)

root.mainloop()